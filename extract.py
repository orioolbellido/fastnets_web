import openpyxl
import glob
import json
import os
from datetime import datetime

project_ids = {
    "Projecte 1": "ofis-sa",
    "Projecte 2": "bar-jb",
    "Projecte 3": "hotel-pilson",
    "Projecte 4": "auditori-vilaneta"
}

results = {}

files = sorted(glob.glob("data/Fastnets/**/*.xlsx", recursive=True))
for path in files:
    if "Temporització" not in path: continue
    
    pid = next((v for k, v in project_ids.items() if k in path), None)
    if not pid: continue
    
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    sections = []
    current_section = {"name": "General", "tasks": []}
    
    # find row with "Ordre"
    start_row = 10
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        if row[1] == "Ordre" or row[2] == "Tasca":
            start_row = i + 2
            break
            
    for row in ws.iter_rows(min_row=start_row, max_row=45, values_only=True):
        if not any(cell is not None for cell in row): continue
        if row[1] == "Dies" or row[2] == "Hores" or str(row[2]) == "None" or not str(row[2]).strip(): continue
        
        tasca = str(row[2]).strip()
        start_dt = row[4]
        end_dt = row[5]
        
        # Check if it's a section header (no dates)
        if not start_dt and not end_dt and "Cablejat" in tasca or "Xarxa" in tasca or "Megafonia" in tasca or "Video" in tasca or "I·luminació" in tasca:
            if len(current_section["tasks"]) > 0:
                sections.append(current_section)
            current_section = {"name": tasca, "tasks": []}
            continue
            
        # Parse dates
        if isinstance(start_dt, datetime): start_str = start_dt.strftime("%Y-%m-%d")
        else: start_str = str(start_dt).split()[0] if start_dt else ""
        
        if isinstance(end_dt, datetime): end_str = end_dt.strftime("%Y-%m-%d")
        else: end_str = str(end_dt).split()[0] if end_dt else ""
        
        if not start_str or start_str == "None": continue
        
        # calculate progress randomly or complete
        t = {
            "id": "t" + str(len(current_section["tasks"])),
            "name": tasca,
            "startDate": start_str,
            "endDate": end_str,
            "progress": 100,
            "status": "Completat"
        }
        current_section["tasks"].append(t)
        
    if len(current_section["tasks"]) > 0:
        sections.append(current_section)
        
    results[pid] = sections

with open("gantt_real.json", "w") as f:
    json.dump(results, f, indent=2)

